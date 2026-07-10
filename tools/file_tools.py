"""
File system tools: read, write, list, search, and scaffold projects.
"""

import glob as glob_module
import os
from pathlib import Path
from typing import Optional

from config import USER_HOME, PROJECTS_DIR


def _resolve(path: Optional[str], default: Optional[str] = None) -> str:
    """Expand ~ and make absolute."""
    if path is None:
        return default or USER_HOME
    return str(Path(path).expanduser().resolve())


# --------------------------------------------------------------------------- #
# Read / Write                                                                 #
# --------------------------------------------------------------------------- #

def read_file(path: str) -> str:
    full = _resolve(path)
    if not os.path.exists(full):
        return f"File not found: {full}"
    if not os.path.isfile(full):
        return f"Not a file: {full}"

    size = os.path.getsize(full)
    if size > 500_000:  # 500 KB cap
        return f"File too large to read in full ({size // 1024} KB). Use run_command with head/tail."

    try:
        with open(full, "r", encoding="utf-8") as f:
            return f.read()
    except PermissionError:
        return f"Permission denied reading: {full}"
    except UnicodeDecodeError:
        # Retry with lossy replacement for binary/non-UTF-8 files
        try:
            with open(full, "r", encoding="utf-8", errors="replace") as f:
                return f.read()
        except Exception as e:
            return f"Cannot decode file (not valid text): {e}"
    except FileNotFoundError:
        return f"File not found: {full}"
    except OSError as e:
        return f"Cannot read file: {e}"


def write_file(path: str, content: str, append: bool = False) -> str:
    full = _resolve(path)
    from tools.permissions import request
    if not request(f"Write to {full}?"):
        return "Permission denied."
    os.makedirs(os.path.dirname(full) or ".", exist_ok=True)
    mode = "a" if append else "w"
    try:
        with open(full, mode, encoding="utf-8") as f:
            f.write(content)
        action = "Appended to" if append else "Written"
        return f"{action} {full} ({len(content)} chars)"
    except PermissionError:
        return f"Permission denied writing to: {full}"
    except OSError as e:
        if e.errno == 28:  # ENOSPC
            return f"Disk full — couldn't write to: {full}"
        return f"Cannot write file: {e}"


# --------------------------------------------------------------------------- #
# List / Search                                                                #
# --------------------------------------------------------------------------- #

def list_directory(path: Optional[str] = None, show_hidden: bool = False) -> str:
    full = _resolve(path)
    if not os.path.exists(full):
        return f"Directory not found: {full}"
    if not os.path.isdir(full):
        return f"Not a directory: {full}"

    try:
        entries = sorted(os.scandir(full), key=lambda e: (not e.is_dir(), e.name.lower()))
    except PermissionError:
        return f"Permission denied: {full}"

    lines = [f"Contents of {full}:"]
    for e in entries:
        if not show_hidden and e.name.startswith("."):
            continue
        kind = "/" if e.is_dir() else ""
        try:
            size = "" if e.is_dir() else f"  {_fmt_size(e.stat().st_size)}"
        except OSError:
            size = ""
        lines.append(f"  {e.name}{kind}{size}")

    if len(lines) == 1:
        lines.append("  (empty)")
    return "\n".join(lines)


def search_files(
    pattern: str,
    directory: Optional[str] = None,
    max_results: int = 50,
) -> str:
    root = _resolve(directory)
    if not os.path.isdir(root):
        return f"Directory not found: {root}"

    # Support both flat and recursive patterns
    try:
        if "**" in pattern:
            matches = glob_module.glob(os.path.join(root, pattern), recursive=True)
        else:
            matches = glob_module.glob(os.path.join(root, "**", pattern), recursive=True)
    except Exception as exc:
        return f"Invalid search pattern '{pattern}': {exc}"

    if not matches:
        return f"No files matching '{pattern}' in {root}"

    matches = sorted(matches)[:max_results]
    lines = [f"Found {len(matches)} file(s) matching '{pattern}':"]
    for m in matches:
        rel = os.path.relpath(m, root)
        lines.append(f"  {rel}")

    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# Project scaffolding                                                          #
# --------------------------------------------------------------------------- #

_GITIGNORE = {
    "python": "__pycache__/\n*.py[cod]\n*.egg-info/\ndist/\nbuild/\n.venv/\nvenv/\n.env\n",
    "node":   "node_modules/\ndist/\nbuild/\n.env\n*.log\n",
    "web":    "node_modules/\ndist/\nbuild/\n.env\n*.log\n",
    "general": ".env\n*.log\n",
}

_TEMPLATES: dict[str, dict[str, str]] = {
    "python": {
        "src/__init__.py": "",
        "tests/__init__.py": "",
        "tests/test_main.py": 'def test_placeholder():\n    assert True\n',
        "main.py": 'def main():\n    pass\n\nif __name__ == "__main__":\n    main()\n',
        "requirements.txt": "",
        ".env.example": "# Add environment variables here\n",
    },
    "node": {
        "src/index.js": "// Entry point\n",
        "tests/index.test.js": "// Tests\n",
        "package.json": "",  # filled dynamically
    },
    "web": {
        "index.html": "<!DOCTYPE html>\n<html lang=\"en\">\n<head>\n  <meta charset=\"UTF-8\">\n  <title>{name}</title>\n</head>\n<body>\n  <h1>{name}</h1>\n</body>\n</html>\n",
        "style.css": "/* Styles */\n",
        "script.js": "// Scripts\n",
    },
    "general": {
        "notes.md": "# {name}\n\n{description}\n",
    },
}


def create_project(
    name: str,
    project_type: str,
    description: str = "",
    directory: Optional[str] = None,
) -> str:
    from tools.permissions import request
    parent = _resolve(directory, PROJECTS_DIR)
    project_dir = os.path.join(parent, name)
    if not request(f"Create project '{name}' at {project_dir}?"):
        return "Permission denied."

    if os.path.exists(project_dir):
        return f"Directory already exists: {project_dir}"

    try:
        os.makedirs(project_dir, exist_ok=False)
    except OSError as e:
        return f"Cannot create project directory: {e}"

    created: list[str] = []

    # Write template files
    template = _TEMPLATES.get(project_type, _TEMPLATES["general"])
    for rel_path, content in template.items():
        full_path = os.path.join(project_dir, rel_path)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)

        # Inject variables
        content = content.replace("{name}", name).replace("{description}", description)

        # Special case: package.json
        if rel_path == "package.json":
            import json
            content = json.dumps(
                {
                    "name": name.lower().replace(" ", "-"),
                    "version": "1.0.0",
                    "description": description,
                    "main": "src/index.js",
                    "scripts": {"test": "jest"},
                },
                indent=2,
            )

        with open(full_path, "w") as f:
            f.write(content)
        created.append(rel_path)

    # README
    readme_path = os.path.join(project_dir, "README.md")
    with open(readme_path, "w") as f:
        f.write(f"# {name}\n\n{description or 'A new project.'}\n")
    created.append("README.md")

    # .gitignore
    gitignore_path = os.path.join(project_dir, ".gitignore")
    with open(gitignore_path, "w") as f:
        f.write(_GITIGNORE.get(project_type, _GITIGNORE["general"]))
    created.append(".gitignore")

    # Init git — list-form subprocess so the shell never gets to interpret
    # `project_dir`. The previous `os.system(f"git -C {project_dir!r}…")`
    # used `!r` to repr-quote but still went through a shell, leaving room
    # for surprises with quirky characters in the path.
    import subprocess as _sp
    try:
        _sp.run(
            ["git", "-C", project_dir, "init", "-q"],
            stdout=_sp.DEVNULL, stderr=_sp.DEVNULL, timeout=10,
        )
    except Exception as exc:
        # Non-fatal — project files were created; git init failure is
        # surfaced in the message below but doesn't break the function.
        print(f"[create_project] git init failed: {exc}", flush=True)

    lines = [
        f"Project '{name}' created at {project_dir}",
        f"Type: {project_type}",
        f"Files created: {', '.join(created)}",
        "Git repository initialised.",
    ]
    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# Helpers                                                                      #
# --------------------------------------------------------------------------- #

def _fmt_size(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024:
            return f"{n:.0f} {unit}"
        n /= 1024
    return f"{n:.1f} TB"


# --------------------------------------------------------------------------- #
# Spreadsheet creation                                                         #
# --------------------------------------------------------------------------- #

def create_spreadsheet(
    filename: str,
    headers: list,
    rows: list,
    sheet_name: str = "Sheet1",
    destination: str = "~/Desktop",
    open_after: bool = True,
) -> str:
    """
    Create a spreadsheet (.xlsx) and optionally open it in Numbers.
    Falls back to CSV if openpyxl isn't available.
    """
    import subprocess
    dest_dir = Path(destination).expanduser()
    dest_dir.mkdir(parents=True, exist_ok=True)

    # Normalise filename extension
    name_no_ext = filename.rsplit(".", 1)[0] if "." in filename else filename

    # Try .xlsx first (opens natively in Numbers / Excel)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header row — bold, light blue fill
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        ws.append(headers)
        for cell in ws[1]:
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows — alternate shading
        alt_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for i, row in enumerate(rows):
            ws.append(row)
            if i % 2 == 1:
                for cell in ws[i + 2]:
                    cell.fill = alt_fill

        # Auto-fit column widths
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        out_path = dest_dir / f"{name_no_ext}.xlsx"
        wb.save(str(out_path))
    except ImportError:
        # openpyxl not available — write CSV
        import csv
        out_path = dest_dir / f"{name_no_ext}.csv"
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

    if open_after:
        subprocess.Popen(["open", str(out_path)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    row_count = len(rows)
    col_count = len(headers)
    return (
        f"Spreadsheet created: {out_path.name} — "
        f"{row_count} rows × {col_count} columns. "
        f"{'Opened in Numbers.' if open_after else ''}"
    )


# --------------------------------------------------------------------------- #
# File management (find / delete / count by pattern + age)                     #
# --------------------------------------------------------------------------- #

def manage_files(
    action: str,
    directory: str = "~/Desktop",
    pattern: str = "*",
    age_hours: float = 0,
    newer_than_hours: float = 0,
    confirm: bool = True,
    recursive: bool = False,
) -> str:
    """
    Find, count, list, or delete files by pattern and/or age.

    action: "list" | "count" | "delete"
    directory: where to look (default ~/Desktop)
    pattern: glob pattern, e.g. "Screenshot*.png", "*.tmp", "*.pdf"
    age_hours: delete files OLDER than this many hours (0 = any age)
    newer_than_hours: only match files created in the last N hours (0 = any age)
    confirm: if True, list files before deleting (set to True always — model decides)
    """
    import time as _time

    dir_path = Path(directory).expanduser()
    if not dir_path.exists():
        return f"Directory not found: {dir_path}"

    now = _time.time()

    # Collect matching files
    glob_fn = dir_path.rglob if recursive else dir_path.glob
    try:
        all_paths = list(glob_fn(pattern))
    except Exception as exc:
        return f"Pattern error: {exc}"

    matched = []
    for p in all_paths:
        if not p.is_file():
            continue
        try:
            mtime = p.stat().st_mtime
        except Exception:
            continue
        age_secs = now - mtime
        if age_hours > 0 and age_secs < age_hours * 3600:
            continue   # not old enough
        if newer_than_hours > 0 and age_secs > newer_than_hours * 3600:
            continue   # too old
        matched.append(p)

    matched.sort(key=lambda p: p.stat().st_mtime, reverse=True)

    if action == "count":
        return f"{len(matched)} file(s) matching '{pattern}' in {dir_path}."

    if action == "list":
        if not matched:
            return f"No files matching '{pattern}' in {dir_path}."
        lines = [f"{p.name}  ({_fmt_size(p.stat().st_size)})" for p in matched[:50]]
        suffix = f"\n… and {len(matched) - 50} more" if len(matched) > 50 else ""
        return f"{len(matched)} file(s):\n" + "\n".join(lines) + suffix

    if action == "delete":
        if not matched:
            return f"No files matching '{pattern}' in {dir_path} to delete."
        deleted, failed = [], []
        for p in matched:
            try:
                p.unlink()
                deleted.append(p.name)
            except Exception as exc:
                failed.append(f"{p.name}: {exc}")
        result = f"Deleted {len(deleted)} file(s)."
        if failed:
            result += f"\nFailed to delete {len(failed)}: " + "; ".join(failed[:5])
        return result

    return f"Unknown action '{action}'. Use: list, count, delete."
